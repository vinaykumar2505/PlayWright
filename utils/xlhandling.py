import pytest


from openpyxl import load_workbook
@pytest.mark.test7
def test_worbook():
    workbook = load_workbook("testdata.xlsx")
    sheet = workbook["Sheet1"]
    value = sheet.cell(row=2, column=1).value
    print(value)

